import os
import glob
import re
import numpy as np
import uuid
import tkinter as tk
import subprocess
import threading
import psutil
from lxml import etree
from pathlib import Path
from openpyxl import load_workbook

# SETUP STEPS:
# Resolution to 1920/1080, set  Export Report Set to Ctrl + W
# If not using MBDVidia, there must be an excel Doc either exported from MBDVidia or having all of the annoations stored in the G column
# and the geometry information stored as a QIF file in the WORKING DOCUMENTS file
# Runnning it as an administrator makes it so the error popup doesn't get sent. However, it doesn't affect anything anyways.
# When tolerancing small parts (dimensions around 0.5 in) the display can get in way of the part.

class GlobalID:
    # Class-level variable to store the counter
    idMax = 0

    @staticmethod
    def increment():
        GlobalID.idMax += 1
        return GlobalID.idMax

    @staticmethod
    def set_id(value):
        GlobalID.idMax = value
    
    @staticmethod
    def get_id():
        id = GlobalID.idMax
        GlobalID.increment()
        return id
    
    @staticmethod
    def peek_id():
        return GlobalID.idMax

# creates a popup that asks the user to input the default tolerances for the part
def get_values_from_popup():
    # Create the main tkinter window
    root = tk.Tk()
    root.withdraw()  # Hide the root window

    # Create a new popup window
    popup = tk.Toplevel()
    popup.title("Enter Default Tolerances")

    # Add a header label
    header = tk.Label(popup, text="Enter the default tolerances found in the title block:", font=("Arial", 12, "bold"))
    header.grid(row=0, column=0, columnspan=2, padx=10, pady=10)

    # Labels and entry fields
    labels = ["X.", ".X", ".XX", ".XXX", ".XXXX"]
    entries = []

    for i, label in enumerate(labels):
        tk.Label(popup, text=label).grid(row=i + 1, column=0, padx=10, pady=5, sticky="w")
        entry = tk.Entry(popup)
        entry.grid(row=i + 1, column=1, padx=10, pady=5)
        entries.append(entry)

    # Variable to store the values
    values = []

    # Function to handle the "Okay" button click or "Enter" key press
    def on_okay(event=None):  # `event` is optional for key bindings
        # Get the values from the text boxes
        nonlocal values
        values = [entry.get() for entry in entries]
        popup.destroy()  # Close the popup window

    # Add the "Okay" button
    tk.Button(popup, text="Okay", command=on_okay).grid(row=len(labels) + 1, column=0, columnspan=2, pady=10)

    # Bind the "Enter" key to the `on_okay` function
    popup.bind("<Return>", on_okay)

    popup.protocol("WM_DELETE_WINDOW", on_okay)

    # Run the tkinter main loop
    popup.wait_window()  # Wait for the popup window to close

    for index, value in enumerate(values):
        values[index] = value.strip('0')
        if (value == ''):
            values[index] = '0'
    # Return the collected values
    return values


exe_finished = threading.Event()
def run_exe():
        exe_folder = os.path.join(os.getcwd(), "resources/AHK Scripts")
        exe_path = os.path.join(exe_folder, "execmacro.exe")
        result = subprocess.run([exe_path])
        exe_finished.set()

def on_popup_close():
    global popup_closed
    popup_closed = True
    root.destroy()


def usingMBD():
    """
    Creates a popup box asking if the user needs to convert files to QIF with MBDVidia.
    Returns True if 'Yes' is pressed, False if 'No' is pressed.
    """
    # Create the main tkinter window
    global root
    root = tk.Tk()
    root.withdraw()  # Hide the root window

    # Create a new popup window
    popup = tk.Toplevel()
    popup.title("MBDVidia Conversion")

    # Add a label with the prompt
    label = tk.Label(popup, text="Do you need to convert your part to a QIF and Excel Document with MBDVidia?", font=("Arial", 10))
    label.pack(padx=10, pady=20)

    # Variable to store the user's choice
    result = tk.BooleanVar(value=False)

    # Function to handle the "Yes" button
    def on_yes():
        result.set(True)
        popup.destroy()

    # Function to handle the "No" button
    def on_no():
        result.set(False)
        popup.destroy()

    # Add "Yes" and "No" buttons
    yes_button = tk.Button(popup, text="Yes", command=on_yes, width=10)
    yes_button.pack(side=tk.LEFT, padx=10, pady=10)

    no_button = tk.Button(popup, text="No", command=on_no, width=10)
    no_button.pack(side=tk.RIGHT, padx=10, pady=10)

    # Center the popup on the screen
    popup.update_idletasks()  # Ensure the window is drawn before calculating dimensions
    screen_width = popup.winfo_screenwidth()
    screen_height = popup.winfo_screenheight()
    popup_width = popup.winfo_reqwidth()
    popup_height = popup.winfo_reqheight()
    x = (screen_width // 2) - (popup_width // 2)
    y = (screen_height // 2) - (popup_height // 2) - 200
    popup.geometry(f"{popup_width}x{popup_height}+{x}+{y}") 

    # Wait for the popup window to close
    popup.wait_window()

    # Return the user's choice
    return result.get()


def getIDMax(root):
    return root.get("idMax")

#given the element and nominal, print out to the terminal
def printElement(element, nominal):
    print(f"{etree.QName(element).localname} of: {nominal}")
    print(etree.tostring(element, pretty_print=True, encoding="unicode"))

def delete_all_files_in_folder(folder_path):
    close_all_instances_of_mbdvidia()
    # Get a list of all files in the folder
    files = glob.glob(os.path.join(folder_path, "*"))

    # Loop through the files and delete them
    for file in files:
        try:
            os.remove(file)
            print(f"Deleted: {file}")
        except Exception as e:
            print(f"Error deleting {file}: {e}")

def close_all_instances_of_mbdvidia():
    """
    Closes all instances of the "MBDVidia" application.
    """
    for process in psutil.process_iter(['name']):
        try:
            # Check if the process name matches "MBDVidia"
            if process.info['name'] and "MBDVidia" in process.info['name']:
                process.terminate()  # Terminate the process
                print(f"Terminated: {process.info['name']} (PID: {process.pid})")
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass


def get_excel_data_as_numpy_array(folder_name):
    """
    Finds the first .xlsx file in the 'temp' folder and extracts its data into a 2D NumPy array.

    Returns:
        np.ndarray: A 2D NumPy array containing the data from the Excel file.
    """
    # Define the path to the 'temp' folder
    temp_folder = os.path.join(os.getcwd(), folder_name)
    print(temp_folder)

    # Find all .xlsx files in the 'temp' folder
    xlsx_files = glob.glob(os.path.join(temp_folder, "*.xlsx"))
    print(xlsx_files)

    # Check if there are any .xlsx files in the folder
    if not xlsx_files:
        raise FileNotFoundError("No .xlsx files found in the 'temp' folder.")

    # Use the first .xlsx file found
    file_path = xlsx_files[0]

    # Load the workbook and select the active sheet
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    # Extract all rows of data from the sheet
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # Convert the data to a 2D NumPy array
    data_array = np.array(data, dtype=str)

    return data_array

# extracts the values from column 'G' (7th column) of the 2D NumPy array
# and returns them as a list
def extract_column_g_from_array(data_array):
    """
    Extracts all values from column 'G' (7th column) of the given 2D NumPy array.

    Args:
        data_array (np.ndarray): A 2D NumPy array containing the Excel data.

    Returns:
        list: A list of values from column 'G'.
    """
    # Ensure the array has enough columns for column G
    if data_array.shape[1] < 7:
        raise IndexError("The data does not contain a 'G' column (7th column).")

    # Extract the 7th column (index 6) and convert it to a Python list
    column_g_values = data_array[:, 6].tolist()
    column_b_values = data_array[:, 1].tolist()

    return column_g_values, column_b_values

def convertQIFtoXML(folder_name):
    # Define the path to the 'temp' folder
    temp_folder = os.path.join(os.getcwd(), folder_name)

    # Find all .qif files in the 'temp' folder
    qif_files = glob.glob(os.path.join(temp_folder, "*.qif"))

    # Check if there are any .qif files in the folder
    if not qif_files:
        print("No .qif files found in the 'temp' folder.")
        xml_files = glob.glob(os.path.join(temp_folder, "*.xml"))
        new_file_path = xml_files[0]

    else:
        # Convert the first .qif file found to XML
        file_path = qif_files[0]
        new_file_path = file_path.replace(".qif", ".xml")
        old_name = Path(file_path)
        new_name = Path(new_file_path)
        old_name.rename(new_name)

    # Read the XML file as a string
    with open(new_file_path, 'r', encoding='utf-8') as file:
        xml_content = file.read()

    # Replace the invalid namespace '##other' with an empty string
    updated_content = xml_content.replace('xmlns="##other"', 'xmlns="http://qifstandards.org/xsd/qif3"')

    # Write the updated content back to the file
    with open(new_file_path, 'w', encoding='utf-8') as file:
        file.write(updated_content)
    print("Invalid namespace '##other' has been removed.")

    return new_file_path

def find_precision(nom, upper_tol, lower_tol):
    # Helper function to calculate precision of a number
    def get_precision(value):
        if '.' in value:
            return len(value.split('.')[1])  # Count digits after the decimal point
        return 0  # No decimal point means precision is 0

    # Calculate precision for the nominal value
    nom_precision = get_precision(nom)

    # Calculate precision for each value in the tolerance list
    up_tol_precision = get_precision(upper_tol)
    low_tol_precision = get_precision(lower_tol)



    # Return the highest precision
    return max(nom_precision, up_tol_precision, low_tol_precision)


# this should return a dictionary of the useful annotations, key is the nominal value, value is a 2 element list,
# first being lower tolerance and second being upper tolerance, the third being the precision, the fourth being the tag
def isolateDiaAnnotes(annotations, default_tol_values, tags):
    #Isolates useful annotations and strips them just as their nominals
    dia_annotes = {}
    for index, annotation in enumerate(annotations):
        if ("DIA" in annotation or "Ø" in annotation) and ("{" not in annotation):
            nom = find_nominal(annotation)
            tol_list, smallest_precision, valid_tol = findTol(annotation, nom, default_tol_values)

            #round each of the values to the correct precision for each
            # annote
            if valid_tol:
                dia_annotes[f"{float(nom):.{smallest_precision}f}"] = tol_list + [smallest_precision] + [tags[index]]
            
    return dia_annotes


def findDefaultUnit(root, ns):
    unit = root.find(".//qif:PrimaryUnits/qif:LinearUnit/qif:UnitName", namespaces=ns).text
    return unit

#the function finds the tolerance from the given string
def findTol(str, nom, default_tol_values):
    precision = 0
    upper_tol = 0
    lower_tol = 0
    valid_tol = True

    if '±' in str:
        match = re.search(r"±\s*(-?\d*\.?\d+)", str)
        upper_tol = lower_tol = match.group(1)
    elif (" " in str and "+" in str and "-" in str):
        # there are spaces within the annotation, this means that it is a unilaterial tolerance
        # Extract the first number after the "+" symbol
        upper_match = re.search(r"\+\s*(-?\d*\.?\d+)", str)
        if upper_match:
            upper_tol = upper_match.group(1)
        else: 
            valid_tol = False

        # Extract the first number after the "-" symbol
        lower_match = re.search(r"-\s*(-?\d*\.?\d+)", str)
        if lower_match:
            lower_tol = lower_match.group(1)
        else: 
            valid_tol = False
    else:
    #if the diameter is by itself - set it to default tolerances
        if '.' in nom:
            precision = len(nom.split('.')[1])
        else:
            precision = 0
        upper_tol = lower_tol = default_tol_values[precision]
    
    smallest_precision = find_precision(nom, upper_tol, lower_tol)

    tol_array = [(float(lower_tol)), (float(upper_tol))]
    tol_array[0] = f"{round(tol_array[0], smallest_precision):.{smallest_precision}f}"
    tol_array[1] = f"{round(tol_array[1], smallest_precision):.{smallest_precision}f}"

    return tol_array, smallest_precision, valid_tol
    

def find_nominal(input_string):
    # Regular expression to match a number (including decimals)
    # Regular Expression (r"-?\d*\.?\d+"):

    # -?: Matches an optional negative sign.
    # \d*: Matches zero or more digits before the decimal point.
    # \.?: Matches an optional decimal point.
    # \d+: Matches one or more digits after the decimal point.

    match = re.search(r"-?\d*\.?\d+", input_string)
    if match:
        return match.group()  # Return the matched number as a string
    return None

def round_to_thousandths(number_str):
    try:
        # Convert the string to a float, round to 3 decimal places, and convert back to a string
        rounded_number = round(float(number_str), 3)
        return str(rounded_number)  # Format to ensure 3 decimal places
    except ValueError:
        return None


# looks for 00, if found truncates after. Looks for 99, if so rounds up after. If neither, just returns raw value
def getDiameter(element, ns, scale):
    core_element = element.find(".//qif:Diameter", namespaces=ns).text
    core_element = str(float(core_element) / float(scale))
    if '00' in core_element:
        print("Found 00 in core element ", core_element)
        integer_part, decimal_part = core_element.split('.', 1)
        second_zero_index = decimal_part.find('0', decimal_part.find('0') + 1)
        if second_zero_index != -1:
            truncated_decimal = decimal_part[:second_zero_index + 1]
            return f"{integer_part}.{truncated_decimal}"
    elif '99' in core_element:
        print("Found 99 in core element")
        integer_part, decimal_part = core_element.split('.', 1)
        second_nine_index = decimal_part.find('9', decimal_part.find('9') + 1)
        if second_nine_index != -1:
            rounded_num = round(float(core_element), (second_nine_index + 1))
            return str(rounded_num)
    else:
        # If no special cases, return the raw value
        return core_element


def getLength(element, ns):
    length = element.find(".//qif:Length", namespaces=ns).text
    length = round_to_thousandths(length)
    return length

def getAxisPoint(element, ns):
    axis_point = element.find(".//qif:AxisPoint", namespaces=ns).text
    axis_point = axis_point.split()
    rounded_axis_point = [round_to_thousandths(coord) for coord in axis_point]
    return " ".join(rounded_axis_point)

def getScaleCoefficient(root, ns):
    scale = root.find(".//qif:ScaleCoefficient", namespaces=ns).text
    return scale

def roundListToThousandths(input_list):
    roundedList = []
    for num in input_list:
        try:
            # Convert to float, round to 3 decimal places, and convert back to string
            rounded_number = round(float(num), 3)
            roundedList.append(f"{rounded_number:.3f}")
        except ValueError:
            # Handle invalid numbers
            roundedList.append(num)
    return roundedList

def getListOfCylinderAnnotations(root, diameter_nominals, scale, cleaned_vals):
    potential_diam_elements = []
    rounded_diameter_nominals = roundListToThousandths(diameter_nominals)
    ns = {"qif": "http://qifstandards.org/xsd/qif3"}
    #Loop through all of the Cylinder23Cores and sort them by matching nominals (there should be 2 for each nominal)
    cylinder_elements = root.findall(".//qif:Cylinder23", namespaces=ns)
    for cylinder_element in cylinder_elements:
        diameter = round_to_thousandths(getDiameter(cylinder_element, ns, scale))
        # Check if the diameter is in the list of diameter nominals
        if diameter in diameter_nominals or diameter in cleaned_vals or round_to_thousandths(diameter) in rounded_diameter_nominals:
            potential_diam_elements.append(cylinder_element)
            print("Added to list: ", diameter)
        else:
            print("Not in list: ", diameter)

    if len(potential_diam_elements) == 0:
        for cylinder_element in cylinder_elements:
            scaleV = round_to_thousandths(getDiameter(cylinder_element, ns, scale))
            # Check if the diameter is in the list of diameter nominals
            if (str(float(scaleV) * 2)) in diameter_nominals:
                potential_diam_elements.append(cylinder_element)

    print("List of potential elements before looping and making sure in pairs of -> ", len(potential_diam_elements))
    for ment in potential_diam_elements:
        print(f"Includes: {ment.get('id')}")

    #Loop through potential elements, make sure that all their
    #   lengths and diameters match for each pair
    while True:
        restart = False
        for i in range(0, len(potential_diam_elements) - 1):
            element = potential_diam_elements[i]
            element_diam = round_to_thousandths(getDiameter(element, ns, scale))
            # secondary_elem_found = False
            for j in range(i + 1, len(potential_diam_elements)):
                element2 = potential_diam_elements[j]
                element2_diam = round_to_thousandths(getDiameter(element2, ns, scale))
                if element_diam == element2_diam:
                    # Check if the lengths and diameters match, rounding to *2* decimal places
                    length1 = element.find(".//qif:Length", namespaces=ns).text
                    length2 = element2.find(".//qif:Length", namespaces=ns).text
                    diam1 = element.find(".//qif:Diameter", namespaces=ns).text
                    diam2 = element2.find(".//qif:Diameter", namespaces=ns).text
                    if not (round(float(length1), 3) == round(float(length2), 3) and round(float(diam1), 2) == round(float(diam2), 2)):
                        restart = True
                        for el in potential_diam_elements[:]:
                            if round_to_thousandths(getDiameter(el, ns, scale)) == element_diam:
                                potential_diam_elements.remove(el)
                        print("Removed elements with diameter:", element_diam)
                        break

            if restart:
                break
        if not restart:
            break
    # If there are 3 elements with the same nominal, remove all of them
    # create a list of 2 element arrays, the first one being the nominal and the second one being the # of that nom
    list_nom = {}
    for el in potential_diam_elements:
        diameter = round_to_thousandths(getDiameter(el, ns, scale))
        if diameter not in list_nom:
            list_nom[diameter] = 1
        else:
            # If the nominal is already in the list, remove all elements with that nominal
            list_nom[diameter] += 1

    #remove elements with same nominal if there are 3 of them
    for el2 in list_nom:
        if list_nom[el2] == 3:
            for el in potential_diam_elements[:]:
                if round_to_thousandths(getDiameter(el, ns, scale)) == el2:
                    potential_diam_elements.remove(el)
            print("B/c of 3 elements found, removed diameter:", el2)

    return potential_diam_elements

def generateCylinderDictionary(element_list, ns, nom_tol_dict):
    # sort all the elements into a dictionary, where the key is the nominal and value is list of elements (1 or 2)
    element_dict = {}
    for k in range(len(element_list)):
        element = element_list[k]
        raw_diameter = getDiameter(element, ns, scale)
        thou_diameter = round_to_thousandths(raw_diameter)

        # round both the element's and all the nom_tol_dict nom to thousandths to compare. Then after finding correct nominal,
        # round to the correct precision for the element and store that as the key-1

        nom_list = list(nom_tol_dict.keys())
        found_nom = ""
        for nom in nom_list:
            if thou_diameter == round_to_thousandths(nom):
                found_nom = nom
                break
        
        # found_nom is the correctly rounded nominal corresponding to element


        # loop through all of the elements in dictionary, if the diameter
        # does not match any of the keys, add it as a new key with "-1"
        any_match = False        
        for index, (key, value) in enumerate(element_dict.items()):
            # if the diameter matches a existing key
            if (found_nom == key.split("-")[0]):
                print("Any match is true!\n")
                any_match = True
                break
        # if there are no matches, then add the new nominal as a key
        if (any_match == False):
            print("Any match is false, so adding new key\n")
            element_dict[found_nom + "-1"] = [element]

        # if there is a match, loop through
        #  all matching nominals comparing the 2 element values. If same, add the element to said list.
        # If loop ends w/ no match, create a new key with -n+1 with that new element
        else:
            instance_number = 0
            found_match2 = False
            for index, (key, value) in enumerate(element_dict.items()):
                # if the diameter matches the key
                if (found_nom == key.split("-")[0]):
                    instance_number += 1
                    # check whether the 2 element values match as well
                    length1 = getLength(element, ns)
                    length2 = getLength(value[0], ns)
                    axis_point1 = getAxisPoint(element, ns)
                    axis_point2 = getAxisPoint(value[0], ns)
                    if (length1 == length2 and axis_point1 == axis_point2):
                        # add the element to the list of elements with that nominal
                        element_dict[key].append(element)
                        found_match2 = True
                        break
            # if for loop finishes and no match is found, create a new key with -n+1
            if (found_match2 == False):
                # create a new key with the same nominal but -n+1
                element_dict[found_nom + "-" + str(instance_number + 1)] = [element]
    return(element_dict)

def generateFaceDictionary(cylinder_dict, root, ns):
    # Loop through the cylinder dictionary and find the corresponding <Face> elements for each cylinder
    # and add to new dictionary. Dict has key as the cylinder ID and value as the <Face> element
    face_dict = {}
    # Loop through all of the cylinder dictionary
    for key, elements in cylinder_dict.items():
        # loop through all of the elements in each key
        for el in elements:
            # Get the ID of the cylinder element
            cylinder_id = el.get("id")
            print("Evaluating cylinder ID:", cylinder_id)
            
            face_elements = root.findall(".//qif:Face", namespaces=ns)
            # Loop through all <Face> elements to find the one with the matching ID
            matching_face_element = None
            for face_element in face_elements:
                # Get the ID of the <Face> element
                face_id = face_element.find(".//qif:Surface/qif:Id", namespaces=ns).text
                # Check if the IDs match
                if face_id == cylinder_id:
                    matching_face_element = face_element
                    break
            if (matching_face_element is not None):
                # Add the <Face> element to the dictionary
                face_dict[cylinder_id] = matching_face_element
                print(f"Found matching <Face> for cylinder ID {cylinder_id}")
            else:
                print(f"No matching <Face> found for cylinder ID {cylinder_id}")
    return face_dict

def createFeatureSection(root, ns, cylinder_dict, n_value):
    """
    Creates the <Features> section and inserts it into the XML tree.
    """
    # Check if the <Features> section already exists
    existing_features = root.find(f".//qif:Features", namespaces=ns)
    if existing_features is not None:
        print("<Features> section already exists. Deleting the existing section.")
        # Find the parent of the <Features> section and remove it
        parent = existing_features.getparent()
        parent.remove(existing_features)

    # Create a new <Features> element
    features_section = etree.Element("{http://qifstandards.org/xsd/qif3}Features")

    # Add a <text> element with the content "hello"
    etree.SubElement(features_section, "{http://qifstandards.org/xsd/qif3}FeatureDefinitions", n=f"{n_value}")
    etree.SubElement(features_section, "{http://qifstandards.org/xsd/qif3}FeatureNominals", n=f"{n_value}")
    etree.SubElement(features_section, "{http://qifstandards.org/xsd/qif3}FeatureItems", n=f"{n_value}")
    
    product_element = root.find(f".//qif:Product", namespaces=ns)

    if product_element is not None:
        # Find the parent of <Product>
        parent = product_element.getparent()

        # Get the index of <Product> in its parent's children
        product_index = parent.index(product_element)

        # Insert the <Features> section directly after <Product>
        parent.insert((product_index + 1), features_section)
        print("<Features> section inserted successfully.")
    else:
        print("<Product> element not found. <Features> section not inserted.")
    
    setNValuesForFeatures(root, namespace, cylinder_dict)


def writeXMLtoFile(root, import_folder, export_folder):
    qif_files = glob.glob(os.path.join(import_folder, "*.xml"))
    if not qif_files:
        raise FileNotFoundError("NO QIF FILES FOUND IN FOLDER")
    
    base_name = os.path.splitext(os.path.basename(qif_files[0]))[0]

    annotated_file_name = f"{base_name} - annotated.qif"
    annotated_file_path = os.path.join(export_folder, annotated_file_name)

    tree1 = etree.ElementTree(root)
    tree1.write(
        annotated_file_path,
        encoding="UTF-8",
        xml_declaration=True,
        pretty_print=True
    )

    return annotated_file_path

def generateUUID():
    return str(uuid.uuid4())

def updateNomTol(cylinder_dict, nom_tol_dict):
    new_nom_tol_dict = {}
    print(f"Cylinder dict: {cylinder_dict.keys()}")
    print(f"Nomtol dict: {nom_tol_dict.keys()}")
    for cyl_nom in cylinder_dict:
        print("Evaluating:", cyl_nom)
        nominal = cyl_nom.split('-')[0]
        new_nom_tol_dict[nominal] = nom_tol_dict[nominal]
    return new_nom_tol_dict



def setNValuesForFeatures(root, ns, cylinder_dict):
    # set correct featureDefinitions
    feature_definitions = root.find(".//qif:FeatureDefinitions", namespaces=ns)
    if feature_definitions is not None:
        feature_definitions.set("n", str(len(cylinder_dict)))

    # set correct FeatureNominals
    feature_nominals = root.find(".//qif:FeatureNominals", namespaces=ns)
    if feature_nominals is not None:
        feature_nominals.set("n", str(len(cylinder_dict)))

    # set correct FeatureItems
    feature_items = root.find(".//qif:FeatureNominals", namespaces=ns)
    if feature_items is not None:
        feature_items.set("n", str(len(cylinder_dict)))

def createFeatureNominals(ns, cylinder_list, face_dict):
    # loop through all nominal values and create respective CylinderFeatureNominal from that value
        cylinder_nom_elem = etree.Element("{http://qifstandards.org/xsd/qif3}CylinderFeatureNominal", id=f"{GlobalID.get_id()}")

        #Add the known subElements
        attrib_elem = etree.SubElement(cylinder_nom_elem, "{http://qifstandards.org/xsd/qif3}Attributes", n="1")
        etree.SubElement(attrib_elem, "{http://qifstandards.org/xsd/qif3}AttributeStr", name="_3dv.TrueGeomAlgo", value="Y14.5-2009")
        
        #create the name
        name_elem = etree.SubElement(cylinder_nom_elem, "{http://qifstandards.org/xsd/qif3}Name")
        name_elem.text = f'{cylinder_nom_elem.get("id")}'

        #create the uuid
        uuid_elem = etree.SubElement(cylinder_nom_elem, "{http://qifstandards.org/xsd/qif3}UUID")
        uuid_elem.text = generateUUID()

        #create featuredefinitionID
        feat_def_elem = etree.SubElement(cylinder_nom_elem, "{http://qifstandards.org/xsd/qif3}FeatureDefinitionId")
        feat_def_elem.text = str(GlobalID.peek_id()) 

        #createEntityInternalIds
        # loop through each face that corresponds to each item of the cylinder_list (in this case 2 faces)
        # get those face's ID's and set them as the EntityInteralIds
        entity_internal_ids_elem = etree.SubElement(cylinder_nom_elem, "{http://qifstandards.org/xsd/qif3}EntityInternalIds", n=f"{len(cylinder_list)}")
        #for each cylinder
        for cylinder in cylinder_list:
            #loop through the face_dict that corresponds
            cyl_id = cylinder.get("id")
            face_id = (face_dict[cyl_id]).get("id")
            id_elem = etree.SubElement(entity_internal_ids_elem, "{http://qifstandards.org/xsd/qif3}Id")
            id_elem.text = face_id
        
        #create SubstituteFeatureAlgorithm
        sub_feat_elem = etree.SubElement(cylinder_nom_elem, "{http://qifstandards.org/xsd/qif3}SubstituteFeatureAlgorithm")
        sub_feat_enum_elem = etree.SubElement(sub_feat_elem, "{http://qifstandards.org/xsd/qif3}SubstituteFeatureAlgorithmEnum")
        sub_feat_enum_elem.text = "LEASTSQUARES"

        #create Axis
        axis_elem = etree.SubElement(cylinder_nom_elem, "{http://qifstandards.org/xsd/qif3}Axis")
        axis_point_text = cylinder_list[0].find(".//qif:AxisPoint", namespaces=ns).text
        direction_text = cylinder_list[0].find(".//qif:Direction", namespaces=ns).text

        axis_point_elem = etree.SubElement(axis_elem, "{http://qifstandards.org/xsd/qif3}AxisPoint")
        axis_point_elem.text = axis_point_text
        direction_elem = etree.SubElement(axis_elem, "{http://qifstandards.org/xsd/qif3}Direction")
        direction_elem.text = direction_text
        return cylinder_nom_elem

def createFeatureDefinition(ns, cylinder_element, nominal_value):    
    feature_elem = etree.Element("{http://qifstandards.org/xsd/qif3}CylinderFeatureDefinition", id=f'{GlobalID.get_id()}')

    # SETTING ALL INTERNALEXTERNAL tags to INTERNAL by default!
    int_ex_elem = etree.SubElement(feature_elem, "{http://qifstandards.org/xsd/qif3}InternalExternal")
    int_ex_elem.text = "INTERNAL"

    #set diameter
    diam_elem = etree.SubElement(feature_elem, "{http://qifstandards.org/xsd/qif3}Diameter")
    diam_elem.text = nominal_value.split("-")[0]

    #set length
    length_elem = etree.SubElement(feature_elem, "{http://qifstandards.org/xsd/qif3}Length")
    cyl_length = cylinder_element.find(".//qif:Length", namespaces=ns).text
    length_elem.text = cyl_length

    #set bottom
    bottom_elem = etree.SubElement(feature_elem, "{http://qifstandards.org/xsd/qif3}Bottom")
    buttom_enum_elem = etree.SubElement(bottom_elem, "{http://qifstandards.org/xsd/qif3}BottomEnum")
    buttom_enum_elem.text = "THROUGH"

    return feature_elem


def createFeatureItem(ns, feature_nominal):
    cyl_feat_item_elem = etree.Element("{http://qifstandards.org/xsd/qif3}CylinderFeatureItem", id=f'{GlobalID.get_id()}')

    #create FeatureNominalId
    feat_nom_elem = etree.SubElement(cyl_feat_item_elem, "{http://qifstandards.org/xsd/qif3}FeatureNominalId")
    feat_nom_elem.text = feature_nominal.get("id")

    #create FeatureName
    feat_name_elem = etree.SubElement(cyl_feat_item_elem, "{http://qifstandards.org/xsd/qif3}FeatureName")
    feat_name_elem.text = f"CylinderItem{cyl_feat_item_elem.get('id')}"

    #create DeterminationMode
    det_mode_elem = etree.SubElement(cyl_feat_item_elem, "{http://qifstandards.org/xsd/qif3}DeterminationMode")
    etree.SubElement(det_mode_elem, "{http://qifstandards.org/xsd/qif3}Checked")

    return cyl_feat_item_elem

def createCharacteristicSection(root, ns, n_value):
    # Check if the <Characteristics> section already exists
    existing_characteristics = root.find(f".//qif:Characteristics", namespaces=ns)
    if existing_characteristics is not None:
        print("<Characteristic> section already exists. Deleting the existing section.")
        # Find the parent of the <Features> section and remove it
        parent = existing_characteristics.getparent()
        parent.remove(existing_characteristics)

    # Create a new <Characteristics> element
    characteristics_section = etree.Element("{http://qifstandards.org/xsd/qif3}Characteristics")

    # find the actual <Standard> from the file to put into the ID
    stand_id = root.find(".//qif:Standard", namespaces=ns).get("id")

    # create <FormalStandardID>
    form_stand_elem = etree.SubElement(characteristics_section, "{http://qifstandards.org/xsd/qif3}FormalStandardId")
    form_stand_elem.text = stand_id

    # create <CharacteristicDefinitions>
    char_def_elem = etree.SubElement(characteristics_section, "{http://qifstandards.org/xsd/qif3}CharacteristicDefinitions", n=f"{n_value}")
    # create <CharacteristicNominals>
    char_nom_elem = etree.SubElement(characteristics_section, "{http://qifstandards.org/xsd/qif3}CharacteristicNominals", n=f"{n_value}")
    # create <CharacteristicItems>    
    char_item_elem = etree.SubElement(characteristics_section, "{http://qifstandards.org/xsd/qif3}CharacteristicItems", n=f"{n_value}")

    feature_element = root.find(f".//qif:Features", namespaces=ns)

    if feature_element is not None:
        # Find the parent of <Char>
        parent = feature_element.getparent()

        # Get the index of <char> in its parent's children
        feat_index = parent.index(feature_element)

        # Insert the <Features> section directly after <char>
        parent.insert(feat_index + 1, characteristics_section)
        print("<Characteristics> section inserted successfully.")
    else:
        print("<Features> element not found. <Characteristics> section not inserted.")

def createCharacteristicDefinition(ns, nominal_value, nom_tol_dict):
    #create <CharacteristicDefinition> element
    char_def_elem = etree.Element("{http://qifstandards.org/xsd/qif3}DiameterCharacteristicDefinition", id=f'{GlobalID.get_id()}')

    #create <CharacteristicDesignator> element
    char_designator_elem = etree.SubElement(char_def_elem, "{http://qifstandards.org/xsd/qif3}CharacteristicDesignator")
    etree.SubElement(char_designator_elem, "{http://qifstandards.org/xsd/qif3}Designator")
    uuid_elem = etree.SubElement(char_designator_elem, "{http://qifstandards.org/xsd/qif3}UUID")
    uuid_elem.text = generateUUID()

    #create <FreeState>
    free_state_elem = etree.SubElement(char_def_elem, "{http://qifstandards.org/xsd/qif3}FreeState")
    free_state_elem.text = "false"

    #create <StatisticalCharacteristic>
    stat_char_elem = etree.SubElement(char_def_elem, "{http://qifstandards.org/xsd/qif3}StatisticalCharacteristic")
    stat_char_elem.text = "false"

    #create <UnitedOrContiuousFeature>
    unit_cont_elem = etree.SubElement(char_def_elem, "{http://qifstandards.org/xsd/qif3}UnitedOrContinuousFeature")
    unit_cont_elem.text = "false"

    #create <Tolerance> section
    tolerance_elem = etree.SubElement(char_def_elem, "{http://qifstandards.org/xsd/qif3}Tolerance")
    #get the min and max tolerances from the nom_tol_dict
    key = nominal_value.split('-')[0]
    lower_tol = nom_tol_dict[key][0]
    upper_tol = nom_tol_dict[key][1]
    round_to_precision = nom_tol_dict[key][2]

    max_tol_elem = etree.SubElement(tolerance_elem, "{http://qifstandards.org/xsd/qif3}MaxValue", decimalPlaces=f"{round_to_precision}", linearUnit=f"{findDefaultUnit(root, ns)}")
    max_tol_elem.text = f"{float(upper_tol):.{round_to_precision}f}"

    min_tol_elem = etree.SubElement(tolerance_elem, "{http://qifstandards.org/xsd/qif3}MinValue", decimalPlaces=f"{round_to_precision}", linearUnit=f"{findDefaultUnit(root, ns)}")
    min_tol_elem.text = f"{float(lower_tol)* -1:.{round_to_precision}f}"
    
    def_as_limit_elem = etree.SubElement(tolerance_elem, "{http://qifstandards.org/xsd/qif3}DefinedAsLimit")
    def_as_limit_elem.text = "false"

    return char_def_elem

def createCharacteristicNominal(ns, index, nominal_value, feature_nominal, cylinder_dict, scale, nom_tol_dict):
    
    #create <CharacteristicNominal> element
    char_nom_elem = etree.Element("{http://qifstandards.org/xsd/qif3}DiameterCharacteristicNominal", id=f'{GlobalID.get_id()}')
    
    #create <Attributes> element
    attrib_elem = etree.SubElement(char_nom_elem, "{http://qifstandards.org/xsd/qif3}Attributes", n="2")
    attrib_user_elem = etree.SubElement(attrib_elem, "{http://qifstandards.org/xsd/qif3}AttributeUser", name="_3dv.PmiReportableAdvanced", nameUserAttribute="AttributePmiReportable")
    user_data_XML_elem = etree.SubElement(attrib_user_elem, "{http://qifstandards.org/xsd/qif3}UserDataXML")
    reportable_states_elem = etree.SubElement(user_data_XML_elem, "{http://qifstandards.org/xsd/qif3}ReportableStates", xmlns="http://qifstandards.org/xsd/qif3", n="0")

    attrib_str_elem =  etree.SubElement(attrib_elem, "{http://qifstandards.org/xsd/qif3}AttributeStr", name="CAMC", value="No")

    # create <CharacteristicDefinitionID>
    char_def_elem = etree.SubElement(char_nom_elem, "{http://qifstandards.org/xsd/qif3}CharacteristicDefinitionId")
    char_def_elem.text = str(int(GlobalID.peek_id()) - 2)

    #create <FeatureNominalIds>
    feat_nom_elem = etree.SubElement(char_nom_elem, "{http://qifstandards.org/xsd/qif3}FeatureNominalIds", n="1")
    id_elem = etree.SubElement(feat_nom_elem, "{http://qifstandards.org/xsd/qif3}Id")
    id_elem.text = feature_nominal.get("id")
    
    #create <Name>
    name_elem = etree.SubElement(char_nom_elem, "{http://qifstandards.org/xsd/qif3}Name")
    name_elem.text = f"Feature Size {char_nom_elem.get('id')}"

    #create <CharacteristicDesignator>
    char_designator_elem = etree.SubElement(char_nom_elem, "{http://qifstandards.org/xsd/qif3}CharacteristicDesignator")
    desig_elem = etree.SubElement(char_designator_elem, "{http://qifstandards.org/xsd/qif3}Designator")
    raw_nom = nominal_value.split('-')[0]
    tag_num = nom_tol_dict[raw_nom][3]
    desig_elem.text = tag_num
    uuid_elem = etree.SubElement(char_designator_elem, "{http://qifstandards.org/xsd/qif3}UUID")
    uuid_elem.text = generateUUID()

    #create <TargetValue>
    precision = nom_tol_dict[nominal_value.split("-")[0]][2]
    target_value_elem = etree.SubElement(char_nom_elem, "{http://qifstandards.org/xsd/qif3}TargetValue", decimalPlaces=f"{precision}", linearUnit=f"{findDefaultUnit(root, ns)}")
    target_value_elem.text = nominal_value.split("-")[0]

    return char_nom_elem

def createCharacteristicItems(index, feature_item_elem, characteristic_nominal_elem):
    char_item_elem = etree.Element("{http://qifstandards.org/xsd/qif3}DiameterCharacteristicItem", id=f'{GlobalID.get_id()}')

    #create <Name>
    name_elem = etree.SubElement(char_item_elem, "{http://qifstandards.org/xsd/qif3}Name")
    name_elem.text = f"Diameter Item {char_item_elem.get('id')}"
    
    #create <CharacteristicIDesignator>
    char_designator_elem = etree.SubElement(char_item_elem, "{http://qifstandards.org/xsd/qif3}CharacteristicDesignator")
    desig_elem = etree.SubElement(char_designator_elem, "{http://qifstandards.org/xsd/qif3}Designator")
    desig_elem.text = str(index + 1)
    uuid_elem = etree.SubElement(char_designator_elem, "{http://qifstandards.org/xsd/qif3}UUID")
    uuid_elem.text = generateUUID()

    #create <FeatureItemsIds>
    feat_item_elem = etree.SubElement(char_item_elem, "{http://qifstandards.org/xsd/qif3}FeatureItemIds", n="1")
    id_elem = etree.SubElement(feat_item_elem, "{http://qifstandards.org/xsd/qif3}Id")
    id_elem.text = feature_item_elem.get("id")

    #create <CharacteristicNominalId>
    char_nom_elem = etree.SubElement(char_item_elem, "{http://qifstandards.org/xsd/qif3}CharacteristicNominalId")
    char_nom_elem.text = characteristic_nominal_elem.get("id")

    return char_item_elem

def createViewSetSection(root, ns, n_value):
    #check if the <ViewSet> Section already exists
    existing_viewset = root.find(f".//qif:ViewSet", namespaces=ns)
    if existing_viewset is not None:
        print("<ViewSet> section already exists. Deleting the existing section.")
        # Find the parent of the <Features> section and remove it
        parent = existing_viewset.getparent()
        parent.remove(existing_viewset)

    # Create a new <ViewSet> element
    viewset_section = etree.Element("{http://qifstandards.org/xsd/qif3}ViewSet")

    #create the <Sets>
    camera_set_elem = etree.SubElement(viewset_section, "{http://qifstandards.org/xsd/qif3}CameraSet", n="3")
    savedview_set_elem = etree.SubElement(viewset_section, "{http://qifstandards.org/xsd/qif3}SavedViewSet", n="3")
    annotation_view_set_elem = etree.SubElement(viewset_section, "{http://qifstandards.org/xsd/qif3}AnnotationViewSet", n=f"{n_value}")
    
    #create the 3 default <Camera> elements
    camera_sv_dict = {}
    camera_list = []
    savedview_list = []

    for i in range(3):
        camera_elem = etree.SubElement(camera_set_elem, "{http://qifstandards.org/xsd/qif3}Camera", id=f'{GlobalID.get_id()}')
        view_plane_origin_elem = etree.SubElement(camera_elem, "{http://qifstandards.org/xsd/qif3}ViewPlaneOrigin")
        view_plane_origin_elem.text = "0 0 0"
        orientation_elem = etree.SubElement(camera_elem, "{http://qifstandards.org/xsd/qif3}Orientation")
        value_elem = etree.SubElement(orientation_elem, "{http://qifstandards.org/xsd/qif3}Value")

        ratio_elem = etree.SubElement(camera_elem, "{http://qifstandards.org/xsd/qif3}Ratio")
        ratio_elem.text = "1"
        near_elem = etree.SubElement(camera_elem, "{http://qifstandards.org/xsd/qif3}Near")
        near_elem.text = "1"
        far_elem = etree.SubElement(camera_elem, "{http://qifstandards.org/xsd/qif3}Far")
        far_elem.text = "-1"
        height_elem = etree.SubElement(camera_elem, "{http://qifstandards.org/xsd/qif3}Height")
        height_elem.text = "1"
    
    #create the respective values for each of the elements
    camera_set_elem[0].find(".//qif:Value", namespaces=ns).text = "0.707106781186548 -0.707106781186548 0 0"
    camera_set_elem[1].find(".//qif:Value", namespaces=ns).text = "0.707106781186548 0 0 0.707106781186548"
    camera_set_elem[2].find(".//qif:Value", namespaces=ns).text = "0 -0.707106781186548 0 0.707106781186548"

    #assign the camera list to camera_sv_dict with the key "Cameras"
    camera_list = [camera_set_elem[0], camera_set_elem[1], camera_set_elem[2]]
    camera_sv_dict["Cameras"] = camera_list

    
    #create the 3 default <SavedView> elements
    first_view = False
    for i in range(3):
        savedview_elem = etree.SubElement(savedview_set_elem, "{http://qifstandards.org/xsd/qif3}SavedView", id=f'{GlobalID.get_id()}', label=f"Saved View {i}")
        if i == 0:
            first_view = True
        attributes_elem = etree.SubElement(savedview_elem, "{http://qifstandards.org/xsd/qif3}Attributes", n=f"{4 + first_view}")
        etree.SubElement(attributes_elem, "{http://qifstandards.org/xsd/qif3}AttributeI1", name="_3dv.SavedViewBalloonRange", value="0")
        if first_view:
            etree.SubElement(attributes_elem, "{http://qifstandards.org/xsd/qif3}AttributeI1", name="_3dv.SavedViewBalloonOffset", value="1")
        
        etree.SubElement(attributes_elem, "{http://qifstandards.org/xsd/qif3}AttributeI1", name="_3dv.SavedViewRenderStyle", value="0")
        etree.SubElement(attributes_elem, "{http://qifstandards.org/xsd/qif3}AttributeI1", name="_3dv.SavedViewBalloonSchemaMode", value="0")
        etree.SubElement(attributes_elem, "{http://qifstandards.org/xsd/qif3}AttributeI3", name="_3dv.SavedViewBalloonColor", value="222 222 222")

        if first_view:
            #set <ActiveView to true>
            active_view_elem = etree.SubElement(savedview_elem, "{http://qifstandards.org/xsd/qif3}ActiveView")
            active_view_elem.text = "true"
        
        #create <AnnotationVisibleIds>
        annotation_visible_ids_elem = etree.SubElement(savedview_elem, "{http://qifstandards.org/xsd/qif3}AnnotationVisibleIds", n="0")
        
        #create <CameraIds>
        camera_ids_elem = etree.SubElement(savedview_elem, "{http://qifstandards.org/xsd/qif3}CameraIds", n="1")
        cam_id_elem = etree.SubElement(camera_ids_elem, "{http://qifstandards.org/xsd/qif3}Id")
        cam_id_elem.text = camera_list[i].get("id")

        first_view = False
    
    #add to list and then to camera_sv_dict
    savedview_list = [savedview_set_elem[0], savedview_set_elem[1], savedview_set_elem[2]]
    camera_sv_dict["SavedViews"] = savedview_list

    #add savedview to the root
    top_element = root.find(f".//qif:TopologySet", namespaces=ns)
    if top_element is not None:
        # Find the parent of <Char>
        parent = top_element.getparent()

        # Get the index of <char> in its parent's children
        top_index = parent.index(top_element)

        # Insert the <Features> section directly after <char>
        parent.insert(top_index + 1, viewset_section)
        print("<ViewSet> section inserted successfully.")
    else:
        print("<TopologySet> element not found. <ViewSet> section not inserted.")

    return camera_sv_dict

def createVisualizationSetSection(root, ns, n_value):
    vis_set = etree.Element("{http://qifstandards.org/xsd/qif3}VisualizationSet")

    #create fonts
    fonts_elem = etree.SubElement(vis_set, "{http://qifstandards.org/xsd/qif3}Fonts", n="5")
    
    font0_elem = etree.SubElement(fonts_elem, "{http://qifstandards.org/xsd/qif3}Font", index="0")
    font0_name_elem = etree.SubElement(font0_elem, "{http://qifstandards.org/xsd/qif3}Name")
    font0_name_elem.text = 'Y14.5-2018'
    font0_size_elem = etree.SubElement(font0_elem, "{http://qifstandards.org/xsd/qif3}Size")
    font0_size_elem.text = '8'

    font1_elem = etree.SubElement(fonts_elem, "{http://qifstandards.org/xsd/qif3}Font", index="1")
    font1_name_elem = etree.SubElement(font1_elem, "{http://qifstandards.org/xsd/qif3}Name")
    font1_name_elem.text = 'Arial Unicode MS'
    font1_size_elem = etree.SubElement(font1_elem, "{http://qifstandards.org/xsd/qif3}Size")
    font1_size_elem.text = '8'

    font2_elem = etree.SubElement(fonts_elem, "{http://qifstandards.org/xsd/qif3}Font", index="2")
    font2_name_elem = etree.SubElement(font2_elem, "{http://qifstandards.org/xsd/qif3}Name")
    font2_name_elem.text = 'MS Gothic UI'
    font2_size_elem = etree.SubElement(font2_elem, "{http://qifstandards.org/xsd/qif3}Size")
    font2_size_elem.text = '8'

    font3_elem = etree.SubElement(fonts_elem, "{http://qifstandards.org/xsd/qif3}Font", index="3")
    font3_name_elem = etree.SubElement(font3_elem, "{http://qifstandards.org/xsd/qif3}Name")
    font3_name_elem.text = 'Microsoft YaHei UI'
    font3_size_elem = etree.SubElement(font3_elem, "{http://qifstandards.org/xsd/qif3}Size")
    font3_size_elem.text = '8'

    font4_elem = etree.SubElement(fonts_elem, "{http://qifstandards.org/xsd/qif3}Font", index="4")
    font4_name_elem = etree.SubElement(font4_elem, "{http://qifstandards.org/xsd/qif3}Name")
    font4_name_elem.text = 'Arial'
    font4_size_elem = etree.SubElement(font4_elem, "{http://qifstandards.org/xsd/qif3}Size")
    font4_size_elem.text = '13'

    pmi_display_set = etree.SubElement(vis_set, "{http://qifstandards.org/xsd/qif3}PMIDisplaySet", n=f"{n_value}")

    #add vis set to the root
    #add savedview to the root
    part_set = root.find(f".//qif:PartSet", namespaces=ns)
    if part_set is not None:

        # Find the parent of <Char>
        parent = part_set.getparent()

        # Get the index of <char> in its parent's children
        top_index = parent.index(part_set)

        # Insert the <Features> section directly after <char>
        parent.insert(top_index, vis_set)
        print("<VisualizationSet> section inserted successfully.")

    else:
        print("<ViewSet> element not found. <VisualizationSet> section not inserted.")




# Takes a feature_nominal and finds the direction vector, returns an int 0-2 corresponding to the needed saved view, and the direction of the created annotation
def findDirectionInfo(ns, feature_nominal):
    components = [float(x) for x in feature_nominal.find(".//qif:Direction", namespaces=ns).text.split()]
    #round and make components absolute
    float_components = [round(abs(c), 6) for c in components]
    normalize_components = [int(c) for c in float_components]

    vector = " ".join(map(str, normalize_components))
    one_index = 0
    saved_view_num = 0
    for i in range(len(normalize_components)):
        if (normalize_components[i] == 1.0):
            one_index = i
            break
    
    # translate the index to the actual saved view number
    if (one_index == 0):
        saved_view_num = 2
    elif (one_index == 1):
        saved_view_num = 0
    elif (one_index == 2):
        saved_view_num = 1

    # set direction vector as well (should make it so the text is horizontal in the saved view)
    direction = ''
    if (saved_view_num == 0):
        direction = "1 0 0"
    elif (saved_view_num == 1):
        direction = "0 -1 0"
    elif (saved_view_num == 2):
        direction = "0 0 -1"

    # return the vector and the saved view number
    return vector, saved_view_num, direction    

def createAnnotationView(ns, feature_nominal):
    #create <AnnotationView> element
    annotation_view_elem = etree.Element("{http://qifstandards.org/xsd/qif3}AnnotationView", id=f'{GlobalID.get_id()}')
    vector, saved_view_num, direction = findDirectionInfo(ns, feature_nominal)

    #create <Normal>
    normal_elem = etree.SubElement(annotation_view_elem, "{http://qifstandards.org/xsd/qif3}Normal")
    normal_elem.text = vector
    #create <Direction>
    direction_elem = etree.SubElement(annotation_view_elem, "{http://qifstandards.org/xsd/qif3}Direction")
    direction_elem.text = direction

    return annotation_view_elem

def createPMIDisplay(ns, index, annotation_view, cyl_feature_nominal, dia_char_nominal, dia_char_definition, nominal_value, nom_tol_dict):
    pmi_elem = etree.Element("{http://qifstandards.org/xsd/qif3}PMIDisplay")

    # create <attributes>
    attrib_elem = etree.SubElement(pmi_elem, "{http://qifstandards.org/xsd/qif3}Attributes", n="5")
    etree.SubElement(attrib_elem, "{http://qifstandards.org/xsd/qif3}AttributeI1", name="DimensionDisplayMode", value="1")
    etree.SubElement(attrib_elem, "{http://qifstandards.org/xsd/qif3}AttributeI2", name="_3dv.PmiGroupId", value=f"{index + 1} -1")
    etree.SubElement(attrib_elem, "{http://qifstandards.org/xsd/qif3}AttributeI1", name="_3dv.ArrowheadSide", value="1")
    etree.SubElement(attrib_elem, "{http://qifstandards.org/xsd/qif3}AttributeBool", name="_3dv.DisplayInfoCorrectPositions", value="1")
    etree.SubElement(attrib_elem, "{http://qifstandards.org/xsd/qif3}AttributeBool", name="_3dv.TextAboveLeader", value="0")

    #create <Plane>
    plane_elem = etree.SubElement(pmi_elem, "{http://qifstandards.org/xsd/qif3}Plane")
    annote_view_elem = etree.SubElement(plane_elem, "{http://qifstandards.org/xsd/qif3}AnnotationViewId")
    id_elem = etree.SubElement(annote_view_elem, "{http://qifstandards.org/xsd/qif3}Id")
    id_elem.text = f"{annotation_view.get('id')}"

    #origin of the pmi, basically where it is placed in space, in X, Y, Z format
    # should line up with what plane the annotation is on, calculate this somehow? - 
    origin_elem = etree.SubElement(plane_elem, "{http://qifstandards.org/xsd/qif3}Origin")
    origin_elem.text = getAxisPoint(cyl_feature_nominal, ns)

    #create <Texts>
    texts_elem = etree.SubElement(pmi_elem, "{http://qifstandards.org/xsd/qif3}Texts", lineHeight="3.5", fontIndex="1", n="2")
    #create nominal <Text>
    nom_text_elem = etree.SubElement(texts_elem, "{http://qifstandards.org/xsd/qif3}Text")
    nom_data_elem = etree.SubElement(nom_text_elem, "{http://qifstandards.org/xsd/qif3}Data")
    nom_val = dia_char_nominal.find("qif:TargetValue", namespaces=ns).text
    nom_data_elem.text = f"{{Diameter}}{nominal_value.split('-')[0]}"
    nom_XY_elem = etree.SubElement(nom_text_elem, "{http://qifstandards.org/xsd/qif3}XY")
    nom_XY_elem.text = "0 0"

    #create tolerance <Text>
    tol_text_elem = etree.SubElement(texts_elem, "{http://qifstandards.org/xsd/qif3}Text")
    tol_data_elem = etree.SubElement(tol_text_elem, "{http://qifstandards.org/xsd/qif3}Data")
    XY_elem = etree.SubElement(tol_text_elem, "{http://qifstandards.org/xsd/qif3}XY")


    #check if its a bilateral tolerance:
    raw_nom = nominal_value.split('-')[0]
    if (nom_tol_dict[raw_nom][0] == nom_tol_dict[raw_nom][1]):
        tol_val = dia_char_definition.find(".//qif:MaxValue", namespaces=ns).text
        tol_data_elem.text = f"{{PLUS_MINUS}}{tol_val}"
        XY_elem.text = "20 0"

    else:
        tol_data_elem.text = f"+{nom_tol_dict[raw_nom][1]}{{BR}}-{nom_tol_dict[raw_nom][0]}"
        XY_elem.text = "20 3.15"


    #create <LeaderExtend>
    leader_extend_elem = etree.SubElement(pmi_elem, "{http://qifstandards.org/xsd/qif3}LeaderExtend")
    start_pt_elem = etree.SubElement(leader_extend_elem, "{http://qifstandards.org/xsd/qif3}StartPoint")
    start_pt_elem.text = '0 0' 
    end_pt_elem = etree.SubElement(leader_extend_elem, "{http://qifstandards.org/xsd/qif3}EndPoint")
    end_pt_elem.text = '10 10'
    head_form_elem = etree.SubElement(leader_extend_elem, "{http://qifstandards.org/xsd/qif3}HeadForm")
    head_form_elem.text = 'DOT_FILLED'  
    head_height_elem = etree.SubElement(leader_extend_elem, "{http://qifstandards.org/xsd/qif3}HeadHeight")
    head_height_elem.text = '2.8'
    point_extension_elem = etree.SubElement(leader_extend_elem, "{http://qifstandards.org/xsd/qif3}PointExtension")
    point_extension_elem.text = '0 1.57397'

    #create <Reference>
    ref_elem = etree.SubElement(pmi_elem, "{http://qifstandards.org/xsd/qif3}Reference")
    ref_id_elem = etree.SubElement(ref_elem, "{http://qifstandards.org/xsd/qif3}Id")
    ref_id_elem.text = dia_char_nominal.get('id')

    return pmi_elem


#adds the respective Diameter Characteristic Nominal to the <SavedViewSet>
def addToSavedViewSet(root, ns, feature_nominal, dia_char_nom):
    vector, saved_view_num, direction = findDirectionInfo(ns, feature_nominal)
    
    saved_view_set = root.find(".//qif:SavedViewSet", namespaces=ns)
    saved_view = saved_view_set[saved_view_num]

    # update n value
    curr_annote_vis_ids = saved_view.find(".//qif:AnnotationVisibleIds", namespaces=ns)
    curr_n_val = int(curr_annote_vis_ids.get('n'))
    new_n = curr_n_val + 1

    curr_annote_vis_ids.set("n", str(new_n))
    
    # add <AnnotationViewIds>
    new_id_elem = etree.SubElement(curr_annote_vis_ids, "{http://qifstandards.org/xsd/qif3}Id")
    new_id_elem.text = dia_char_nom.get('id')
    
    print(f"Added {dia_char_nom.get('id')} to the Saved Views")
    
    return saved_view

def updateQIFAttributes(root, ns, n_value):
    attribs_elem = root.find(".//qif:Attributes", namespaces=ns)
    if attribs_elem is not None:
        new_attrib_elem = etree.SubElement(attribs_elem, "{http://qifstandards.org/xsd/qif3}AttributeI1", name="_3dv.QIFDocument.FreeBalloon", value=f"{n_value + 1}")
        attribs_elem.set("n", "10")
        print("Updated QIF Attributes!")
    else:
        print("Attrib not found!")

def updateMassPropertyTolerance(root, ns):
    mass_prop_elem = root.find(".//qif:MassPropertyTolerance", namespaces=ns)
    mass_prop_elem.text = '0.0005'
    print("Updated nass property tolerance!")

def updatePartSet(root, ns, n_value, camera_sv_dict):
    part_elem = root.find(".//qif:PartSet/qif:Part", namespaces=ns)

    model_num_elem = part_elem.find(".//qif:ModelNumber", namespaces=ns)
    model_num_index = part_elem.index(model_num_elem)

    feature_nominal_ids_elem = etree.Element("{http://qifstandards.org/xsd/qif3}FeatureNominalIds", n=f"{n_value}")
    char_nominal_ids_elem = etree.Element("{http://qifstandards.org/xsd/qif3}CharacteristicNominalIds", n=f"{n_value}")




    view_ids_elem = etree.Element("{http://qifstandards.org/xsd/qif3}ViewIds", n=f"{6 + n_value}")
    #loop through the camera to get each of their ids and add them as <Id> elements
    for k in range(3):
        id = camera_sv_dict.get("Cameras")[k].get('id')
        id_elem = etree.SubElement(view_ids_elem, "{http://qifstandards.org/xsd/qif3}Id")
        id_elem.text = id
    #loop through each savedview to get their id and add them
    for j in range(3):
        id = camera_sv_dict.get("SavedViews")[j].get('id')
        id_elem = etree.SubElement(view_ids_elem, "{http://qifstandards.org/xsd/qif3}Id")
        id_elem.text = id

    # if FoldersPart exists, then insert viewIds before it
    folders_part_elem = root.find(".//qif:FoldersPart", namespaces=ns)
    if folders_part_elem is not None:
        folders_part_index = part_elem.index(folders_part_elem)
        part_elem.insert(folders_part_index, view_ids_elem)
    # if it doesn't exist, insert view_ids at the end of the part_elem
    else: 
        part_elem.insert(len(part_elem), view_ids_elem)

    part_elem.insert(model_num_index + 1, char_nominal_ids_elem)
    part_elem.insert(model_num_index + 1, feature_nominal_ids_elem)
    print("Updated PartSet successfully for annotations!")    

def setUpForAnnotations(root, ns, cylinder_dict, n_value):
    #sets the N values within this function as well
    createFeatureSection(root, ns, cylinder_dict, n_value)
    createCharacteristicSection(root, ns, n_value)
    camera_sv_dict = createViewSetSection(root, ns, n_value)
    createVisualizationSetSection(root, ns, n_value)

    updateQIFAttributes(root, ns, n_value)
    updateMassPropertyTolerance(root, ns)
    updatePartSet(root, ns, n_value, camera_sv_dict)

    return camera_sv_dict

def updateIDMax(root, ns):
    root.set("idMax", str(GlobalID.peek_id()))
    print("Updated the ID Max after everything is done: curr max ", GlobalID.peek_id())

def remove_duplicate_features(root, ns):
    """
    Removes all instances of <Features> after the first one in the XML tree.

    Args:
        root (etree.Element): The root element of the XML tree.
        ns (dict): The namespace dictionary for the XML.
    """
    # Find all <Features> elements
    features_elements = root.findall(".//qif:Features", namespaces=ns)

    # Keep the first instance and remove the rest
    if len(features_elements) > 1:
        for features_elem in features_elements[1:]:
            parent = features_elem.getparent()
            if parent is not None:
                parent.remove(features_elem)
                print("<Features> element removed.")

    print("Duplicate <Features> elements removed.")


# main function that loops through the cylinder_dict things and calls the other functiosn to create each element
def createAnnotations(root, ns, cylinder_dict, face_dict, nom_tol_dict, scale):
    
    n_value = len(cylinder_dict)
    camera_sv_dict = setUpForAnnotations(root, ns, cylinder_dict, n_value)


    for index, (nominal_value, cylinder_list) in enumerate(cylinder_dict.items()):
        # create feature info
        feature_nominal = createFeatureNominals(ns, cylinder_list, face_dict)
        # printElement(feature_nominal, nominal_value)

        feature_definition = createFeatureDefinition(ns, cylinder_list[0], nominal_value)
        # printElement(feature_definition, nominal_value)

        feature_item = createFeatureItem(ns, feature_nominal)
        # printElement(feature_item, nominal_value)


        # create characteristic info
        characteristic_definition = createCharacteristicDefinition(ns, nominal_value, nom_tol_dict)
        # printElement(characteristic_definition, nominal_value)

        characteristic_nominal = createCharacteristicNominal(ns, index, nominal_value, feature_nominal, cylinder_dict, scale, nom_tol_dict)
        # printElement(characteristic_nominal, nominal_value)

        characteristic_item = createCharacteristicItems(index, feature_item, characteristic_nominal)
        # printElement(characteristic_item, nominal_value)


        # create visual info and update the saved view
        annotation_view = createAnnotationView(ns, feature_nominal)
        # printElement(annotation_view, nominal_value)

        pmi_display = createPMIDisplay(ns, index, annotation_view, feature_nominal, characteristic_nominal, characteristic_definition, nominal_value, nom_tol_dict)
        # printElement(pmi_display, nominal_value)

        updated_saved_view = addToSavedViewSet(root, ns, feature_nominal, characteristic_nominal)
        # printElement(updated_saved_view, nominal_value)
        

        # add all the elements to the root
        root.find(".//qif:FeatureDefinitions", namespaces=ns).append(feature_definition)
        root.find(".//qif:FeatureNominals", namespaces=ns).append(feature_nominal)
        root.find(".//qif:FeatureItems", namespaces=ns).append(feature_item)
        root.find(".//qif:CharacteristicDefinitions", namespaces=ns).append(characteristic_definition)
        root.find(".//qif:CharacteristicNominals", namespaces=ns).append(characteristic_nominal)
        root.find(".//qif:CharacteristicItems", namespaces=ns).append(characteristic_item)

        root.find(".//qif:AnnotationViewSet", namespaces=ns).append(annotation_view)
        root.find(".//qif:PMIDisplaySet", namespaces=ns).append(pmi_display)

        #in <PartSet> add <FeatureNominalIds>, <CharacteristicNominalIds>, <ViewIds (add the annotation view created)>
        feature_nominal_ids = root.find(".//qif:PartSet/qif:Part/qif:FeatureNominalIds", namespaces=ns)
        feature_nominal_id = etree.SubElement(feature_nominal_ids, "{http://qifstandards.org/xsd/qif3}Id")
        feature_nominal_id.text = feature_nominal.get('id')

        characteristic_nominal_ids = root.find(".//qif:PartSet/qif:Part/qif:CharacteristicNominalIds", namespaces=ns)
        characteristic_nominal_id = etree.SubElement(characteristic_nominal_ids, "{http://qifstandards.org/xsd/qif3}Id")
        characteristic_nominal_id.text = characteristic_nominal.get('id')

        view_ids = root.find(".//qif:PartSet/qif:Part/qif:ViewIds", namespaces=ns)
        annotation_view_id = etree.SubElement(view_ids, "{http://qifstandards.org/xsd/qif3}Id")
        annotation_view_id.text = annotation_view.get('id')

        # print(etree.tostring(root.find(".//qif:Part", namespaces=ns), pretty_print=True, encoding="unicode"))

    #update idMax after all annotations are added
    updateIDMax(root, ns)
    remove_duplicate_features(root, ns)

### Mains
try:
    # This should be a pre created folder where the script is located
    import_folder = "INPUT FILES"
    # - Ask for whether actively using MBDVidia or not, if not then skip the AHK script, either way before ask for the default tolerances
    using_MBDVidia = usingMBD()
    default_tol_list = ['1', '0.1', '0.01', '0.001', '0.0001']
    if using_MBDVidia:
        # start exe in separate thread
        threading.Thread(target=run_exe, daemon=True).start()
        
        #show popup box
        default_tol_list = get_values_from_popup()
        exe_finished.wait()
    else:
        default_tol_list = get_values_from_popup()

    print("Entered default tol list: ", default_tol_list)    

    # Get the Excel data as a 2D NumPy array
    data_array = get_excel_data_as_numpy_array(import_folder)
    g_vals, b_vals = extract_column_g_from_array(data_array)

    # this variable is a dictionary where {nominal : [lower_val, upper_val, precision]}
    nom_tol_dict = isolateDiaAnnotes(g_vals, default_tol_list, b_vals)
    
    #dia_iso_vals is a list of all diameter nominal annotes
    dia_iso_vals = list(nom_tol_dict.keys())
    cleaned_vals = [str(float(num)).rstrip('0').rstrip('.') if '.' in num else str(int(num)) for num in dia_iso_vals]
    
    print("nom_tol_dict is ", nom_tol_dict)

    # Convert the QIF file to XML and reformat
    file_name = convertQIFtoXML(import_folder)

    # Generate the tree
    tree = etree.parse(file_name)
    root = tree.getroot()
    namespace = {"qif": "http://qifstandards.org/xsd/qif3"}
    scale = getScaleCoefficient(root, namespace)
    idMax = int(getIDMax(root))
    GlobalID.set_id((idMax + 1))

    # get the list of all the <Cylinder23> elements that match the diameters in dia_iso_vals
    element_list = getListOfCylinderAnnotations(root, dia_iso_vals, scale, cleaned_vals)

    # from the element list sort by nominal + separate repeated features
    cylinder_dict = generateCylinderDictionary(element_list, namespace, nom_tol_dict)
    # for every nominal, use the ID for each element to find the corresponding 2 <Face> elements
    face_dict = generateFaceDictionary(cylinder_dict, root, namespace)
    # update nom_tol_dict to match the cylinder_dict
    nom_tol_dict = updateNomTol(cylinder_dict, nom_tol_dict)

    # With the cylinder_dict and face_dict create the correct FeatureNominals
    createAnnotations(root, namespace, cylinder_dict, face_dict, nom_tol_dict, scale)

    export_folder = os.path.join(os.getcwd(), "OUTPUT FILES")
    annote_file_path = writeXMLtoFile(root, import_folder, export_folder)
    delete_all_files_in_folder(import_folder)

    if (using_MBDVidia):
        os.startfile(annote_file_path)

except FileNotFoundError as e:
    print(e)

    
# Notes:
# - By default, the first annotation in cylinder_dict is the one that is displayed by default
# - In this patch, if there are 3 elements with the same nominal, they get scrapped,
#   no matter if lengths and dimensions match (example is the 2.2 nom on 0205)
# - When cross checking the extracted noms from the Excel and the noms in the XML, it is initially checked
#   exactly, then if nothing is found it will round both to *3* decimal places and check again.
# - When creating FeatureDefinitions, it will always assign <InternalExternal> to Internal, relying
#   on MBDVidia to heal at the end?
# - Scanning the annotations will only look for symmetrical tolerances
# - All instances of xmlns="##other" have been replaced with xmlns="http://qifstandards.org/xsd/qif3"
# - <CharacteristicNominal> ... <ReportableStates> has been set n = 0, relies on MBDVidia to heal during translation
# - <CharacteristicNominal> ... <AttributeI1> has been removed entirely from all annotes
# - Within <SavedView> the XZ plane saved view is set as the default <ActiveView>

# Diffs I ignored: sometimes in <Part> <Attributes> <_3dv.ReportTypeSaved> value changes to 56 on the annotation model, ignored this
#   - entire <Results> elements
#   - <Lengths> in <CyinderFeatureDefinition> are different that their values in the geometry
#   - big rounding errors, such as in <DiameterCharactersticDefinition>
#   - line 3694, <UserDataXML> is completely ignored

# - UNFIXABLE ISSUES:
# MBD makes it so trailing 0's on the tolerance will not be displayed, so must be created manually
# after the fact. If the nominal requires a trailing 0, cannot even be added
# the software doesn't allow it, even when doing it manually it still gives error
# - Lots of errors with how the geometry values are stored in the QIF, leads to many
# values not being recognized

# Potential improvements::
# If the tech drawing has two identical feature but two different tolerances, this will not work
# PMI scales with the scale of the part, so its not too big and covering up everything (this might only be able to be solved with arrwos?
# how do I put in arrows instead of dots as they would be a lot better.
# - based on both the <Origin> and the <StartPoint> being correct, rest doesn't matter

